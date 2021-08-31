# -*- coding: utf-8 -*-

"""

@author: Emanuele Luzzi

Date: 04/2021

"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import datetime as dt
# import sys


class Organizer():
    
    def __init__(self, file_path, logger):
        
        self.file_path = Path(file_path)
        self.data = self.load_excel_file("EPFR Output")
        self.logger = logger
        
    def load_excel_file(self, sheet_name):
        df = pd.read_excel(self.file_path, sheet_name=sheet_name, index_col=0)
        df = df.fillna(0)
        
        return df
    
    def __repr__(self):
        info = f"Organizer({self.file_path})"
        self.logger.info(info)
        return info
    
    def __str__(self):
        info = f"Organizer in use for the following file: {self.file_path}"
        self.logger.info(info)
        return info
    
    """
    
            ################################
    
             Metrics and Descriptive Stats
                                            
            ################################

    
    """
    
    def calc_sums_metrics(self, dataframe, list_of_addends_labels):
        sums_column = pd.Series(index=sorted(set(self.data.index.tolist()))).fillna(0)
        for addend in list_of_addends_labels:
            to_add_series = dataframe[addend].fillna(0)
            sums_column += to_add_series
        
        return sums_column
    
    """
    
            ################################
    
                Data Matrices Operations
                                            
            ################################

    
    """
    
    def calc_average_between_dataframes_sheets(self, list_of_dataframe_sheets):
        n = len(list_of_dataframe_sheets)
        list_of_dataframes = [self.load_excel_file(sheet).fillna(0) for sheet in list_of_dataframe_sheets]
        start_df = pd.DataFrame(index=list_of_dataframes[0].index, columns = list_of_dataframes[0].columns).fillna(0)
        for df in list_of_dataframes:
            sum_df = start_df.add(df)
            start_df = sum_df
        
        avg_df = sum_df/n
 
        return avg_df
    
    def calc_ratio_between_dataframes_sheets(self, duple_of_dataframe_sheets):
        duple_of_dataframes = [self.load_excel_file(sheet).fillna(0) for sheet in duple_of_dataframe_sheets]
        ratio_df = duple_of_dataframes[0].divide(duple_of_dataframes[1])
        
        return ratio_df
    
    def calc_rebased_dataframe_sheet(self, df_to_rebase, dd_mm_yy, marginal_return_label):
        df_to_rebase = self.load_excel_file(df_to_rebase)
        starting_date = dt.datetime.strptime(dd_mm_yy, "%d.%m.%y")
        marginal_return_dataframe = self.load_excel_file(marginal_return_label)
        rebased_dataframe = pd.DataFrame(index=df_to_rebase.loc[df_to_rebase.index >= starting_date].index, 
                                         columns=df_to_rebase.columns).fillna(0)
        rebased_dataframe.loc[starting_date] = [100] * len(df_to_rebase.columns)
        
        for date in rebased_dataframe.index:
            for column in rebased_dataframe.columns:
                if date != starting_date:
                    marginal_return = marginal_return_dataframe.loc[date, column]
                    index_list = rebased_dataframe.index.tolist()
                    current_position_of_date = index_list.index(date)
                    previous_date = index_list[current_position_of_date - 1]
                    
                    rebased_dataframe.loc[date, column] = (rebased_dataframe.loc[previous_date, column] 
                                                           * (1 + marginal_return) + marginal_return)
                
        return rebased_dataframe
    
    """
    
            ################################
    
                 Data Cleansing Methods
                                            
            ################################

    
    """
        
    def add_dataframe_sheet(self, dataframe, sheet_name):
        xlsx_file = load_workbook(self.file_path)
        writer = pd.ExcelWriter(self.file_path, engine = 'openpyxl')
        writer.book = xlsx_file
        
        try:
            xlsx_file.remove(xlsx_file[sheet_name])
        except Exception as e:
            print(f"Creating a new sheet, since {e}")
        finally:
            dataframe.to_excel(writer, sheet_name = sheet_name)
        
        self.logger.info(f"{sheet_name} added to the sheets")
        
        writer.save()
        writer.close()
       
        
    """
    
    ###############################################################################################
    
                                            Ad hoc toolbox
                                            
    ###############################################################################################

    
    """
    
    def name_for_country_and_filter(self, country, filt):
        country_names = country.split(" ")
        country_name = country_names[0] + " " + country_names[1][0:3]
        filt_names = filt.split(" ")
        filt_name = "".join([word[0] for word in filt_names]).upper()
        
        if country_names[1][0:2] == "ex" or country_names[1][0:2] == "Ex":
            country_name = country_names[0] + " " + country_names[1]
            
        final_name = country_name + " " + filt_name
        
        return final_name
    
    def list_of_names_for_countries_and_filters(self):
        all_countries_in_asset_class = set(self.data["Asset Class"].tolist())
        all_types_in_filters = set(self.data["Filters"].tolist())
        
        name_list = []
        for country in all_countries_in_asset_class:
            for filt in all_types_in_filters:
                name_list.append(self.name_for_country_and_filter(country, filt))
        
        return name_list
    
    def europe_definition(self, members, except_what = []):
        
        ex_countries = []
        for country in self.list_of_names_for_countries_and_filters():
            if country.split("-")[0] in except_what:
                ex_countries.append(country)
        
        europe_members = []
        for member in members:
            for country_and_filter in self.list_of_names_for_countries_and_filters():
                if (member == country_and_filter.split("-")[0] 
                or (country_and_filter.split(" ")[0] == "Europe" and country_and_filter not in europe_members)):
                    europe_members.append(country_and_filter)
            
        europe = [country for country in europe_members if country not in ex_countries]
        
        return europe

    def usa_definition(self, members):
        usa_members = []
        for member in members:
            for country_and_filter in self.list_of_names_for_countries_and_filters():
                if member == country_and_filter.split("-")[0]:
                    usa_members.append(country_and_filter)
            
        usa = [country for country in usa_members]
        
        return usa
    
    def create_dataframe_for_every_country_and_category(self):
        all_countries_in_asset_class = set(self.data["Asset Class"].tolist())
        all_types_in_filters = set(self.data["Filters"].tolist())
        xlsx_file = load_workbook(self.file_path)
        writer = pd.ExcelWriter(self.file_path, engine = 'openpyxl')
        writer.book = xlsx_file
        
        for country in all_countries_in_asset_class:
            for filt in all_types_in_filters:
                sheet_name = self.name_for_country_and_filter(country, filt)
                
                try:
                    xlsx_file.remove(xlsx_file[sheet_name])
                except Exception as e:
                    print(f"Creating a new sheet, since {e}")
                finally:
                    country_and_filt_data = self.data[(self.data["Asset Class"] == country) 
                                                        & (self.data["Filters"] == filt)]
                    country_and_filt_data.to_excel(writer, sheet_name = sheet_name)
                    
                self.logger.info(f"{sheet_name} dataframe created and added to the Excel file.")
                
        writer.save()
        writer.close()
               
    def create_dataframe_for_every_variable_country_and_category(self, europe_members, usa_members):
        all_countries_in_asset_class = set(self.data["Asset Class"].tolist())
        all_types_in_filters = set(self.data["Filters"].tolist())
        variables = [var for var in self.data.columns if var != "Asset Class" and var != "Filters"]
        xlsx_file = load_workbook(self.file_path)
        writer = pd.ExcelWriter(self.file_path, engine = 'openpyxl')
        writer.book = xlsx_file
        
        for var in variables:     
            variable_dataframe = pd.DataFrame(index=sorted(set(self.data.index.tolist())))
            for country in all_countries_in_asset_class:
                for filt in all_types_in_filters:
                    col_name = self.name_for_country_and_filter(country, filt)
                    country_and_filt_data = self.data[(self.data["Asset Class"] == country) 
                                                        & (self.data["Filters"] == filt)]
                    variable_dataframe[col_name] = pd.Series(country_and_filt_data[var]).fillna(0)
                
            sheet_name = var
        
            self.logger.info(f"Specific dataframe for {var} created and added to the Excel file.")
            
            variable_dataframe["Europe"] = self.calc_sums_metrics(variable_dataframe, 
                              self.europe_definition(europe_members))
                        
            variable_dataframe["Europe Ex-IT"] = self.calc_sums_metrics(variable_dataframe, 
                              self.europe_definition(europe_members, "Italy"))
                        
            variable_dataframe["Europe Ex-DE"] = self.calc_sums_metrics(variable_dataframe, 
                              self.europe_definition(europe_members, "Germany"))
                        
            variable_dataframe["Europe Ex-GR"] = self.calc_sums_metrics(variable_dataframe, 
                              self.europe_definition(europe_members, "Greece"))

            variable_dataframe["Overall"] = self.calc_sums_metrics(variable_dataframe, 
                              self.europe_definition(europe_members) + self.usa_definition(usa_members))
            
            try:            
                variable_dataframe["Europe CIG"] = self.calc_sums_metrics(variable_dataframe, 
                                  ["Germany-Western Eur CIG", 
                                   "Europe Reg CIG", 
                                   "Greece-Western Eur CIG", 
                                   "Italy-Western Eur CIG",
                                   "Europe ex-UK CIG"])
                
                variable_dataframe["Europe CHY"] = self.calc_sums_metrics(variable_dataframe, 
                                  ["Germany-Western Eur CHY", 
                                   "Europe Reg CHY", 
                                   "Greece-Western Eur CHY", 
                                   "Italy-Western Eur CHY",
                                   "Europe ex-UK CHY"])
            except KeyError:
                pass
            
            self.logger.info(f"European metrics for {var} created and added to the Sheet.")
            
            try:
                xlsx_file.remove(xlsx_file[sheet_name])
            except Exception as e:
                print(f"Creating a new sheet, since {e}")
            finally:
                variable_dataframe.to_excel(writer, sheet_name = sheet_name)
            
        writer.save()
        writer.close()

    
    
    
    
